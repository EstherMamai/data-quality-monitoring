# scripts/report_generator.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

def style_excel_report(file_path):
    """Apply formatting to the generated Excel report."""
    wb = load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # --- Style header ---
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Auto-size columns ---
        for column in ws.columns:
            max_length = 0
            col = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col].width = max_length + 2

    wb.save(file_path)
    print(f"✅ Report formatting applied successfully: {file_path}")


def create_report(missing_values, invalid_emails, invalid_dates, output_path="../reports/data_quality_report.xlsx"):
    """Create and style the Excel report."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(missing_values, columns=["MissingCount"]).to_excel(writer, sheet_name="Missing Values")
        invalid_emails.to_excel(writer, sheet_name="Invalid Emails", index=False)
        invalid_dates.to_excel(writer, sheet_name="Invalid Dates", index=False)

    style_excel_report(output_path)
